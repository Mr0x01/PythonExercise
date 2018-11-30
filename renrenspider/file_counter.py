import os
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

path = input("指定目录：")
row = 1
for friend in os.listdir(path):
    # 好友信息
    if os.path.isdir(path+"\\"+friend):
        for album in os.listdir(path+"\\"+friend):
                # 相册信息
                ws.cell(row,1,friend)
                ws.cell(row,2,album)
                img_amount = len(os.listdir(path+"\\"+friend+"\\"+album))
                ws.cell(row,3,img_amount)
                row+=1
wb.save(path+"\\info.xlsx")

# fnd_amount(path)
